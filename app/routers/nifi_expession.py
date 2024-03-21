import json
from fastapi import Header, APIRouter,Body,File,UploadFile,HTTPException
from fastapi.responses import JSONResponse
from typing import List,Union,Optional,Annotated 
import pandas as pd
from hashlib import sha256
from datetime import datetime
import requests
import io 
from openpyxl import load_workbook
from unidecode import unidecode
import subprocess
import base64
import polars as pl
import google.cloud.bigquery as bigquery
from google.oauth2 import service_account
import os

def map_csharp_to_python_format(csharp_format: str) -> str:
    # Define the mapping of C# format codes to Python format codes
    mapping = {
        "d": "%d",
        "dd": "%d",
        "M": "%m",
        "MM": "%m",
        "MMM": "%b",
        "MMMM": "%B",
        "y": "%y",
        "yy": "%y",
        "yyyy": "%Y",
        "h": "%I",
        "hh": "%I",
        "H": "%H",
        "HH": "%H",
        "m": "%M",
        "mm": "%M",
        "s": "%S",
        "ss": "%S",
        "f": "%f",
        "ff": "%f",
        "fff": "%f",
        "t": "%p",
        "tt": "%p"
    }
    
    # Map each C# format code to the corresponding Python format code
    python_format = ""
    i = 0
    while i < len(csharp_format):
        if csharp_format[i:i+4] in mapping:
            python_format += mapping[csharp_format[i:i+4]]
            i += 4
        elif csharp_format[i:i+2] in mapping:
            python_format += mapping[csharp_format[i:i+2]]
            i += 2
        elif csharp_format[i] in mapping:
            python_format += mapping[csharp_format[i]]
            i += 1
        else:
            python_format += csharp_format[i]
            i += 1
    return python_format



class BearerAuth(requests.auth.AuthBase):
    def __init__(self, token):
        self.token = token
    def __call__(self, r):
        r.headers["authorization"] = "Bearer " + self.token
        return r

router = APIRouter(prefix="/nifiExpression",tags=["Nifi Expression"])
@router.post('/hashDatalog')
async def hash_datalog(data: List[dict]=Body(...)):
    for i in range(len(data)):
        data[i]['DataLog']=json.dumps(data[i]['DataLog'])
    return(data)
@router.post('/loadDatalog')
async def hash_datalog(data: List[dict]=Body(...)):
    for i in range(len(data)):
        data[i]['DataLog']=json.loads(data[i]['DataLog'])
    return(data)
@router.post('/parseProductAttribute')
async def parseProductAttribute(data: List[dict]=Body(...)):
    df = pd.DataFrame(data)
    data =[]
    product_id=[]
    for index, row in df.iterrows():    
        
        if row['attribute_values'] is not None:
            att = json.loads(row['attribute_values'])
            
            for type1 in att:
                for valuesDict in att[type1]:
                    for type2 in valuesDict:
                        dataRow = dict()
                        dataRow['product_id']= row['product_id']
                        dataRow['product_type']= row['product_type']
                        dataRow['Type1'] = type1
                        dataRow['Type2'] = type2
                        dataRow['Value'] = valuesDict[type2]
                        dataRow['product_name'] = row['product_name']
                        dataRow['product_handle'] = row ['product_handle']
                        dataRow['created_at'] = row['created_at']
                        dataRow['updated_at']=row['updated_at']
                        dataRow['IdRawLog']=row['id']
                        dataRow['id'] = sha256((row['product_id']+ type1 + type2 + valuesDict[type2] + str(datetime.now())).encode('utf-8')).hexdigest()
                        data.append(dataRow)
                        if row['product_id'] not in product_id:
                            product_id.append(row['product_id'])
                       
    return({"product_attribute":data,"product_id": ','.join(str(id) for id in product_id)})

@router.post('/get_onedrive_excel')
async def get_onedrive_excel(data: dict = Body(...)):
    if(data['isShareFile']=='false'):
        response = requests.get('https://graph.microsoft.com/v1.0/me/drive/items/'+data['workbook_id']+'/content', 
            auth=BearerAuth(data['access_token']))
    else: 
        response = requests.get('https://graph.microsoft.com/v1.0/shares/'+data['workbook_id']+'/driveItem/content', 
            auth=BearerAuth(data['access_token']))
    

    bytes_file_obj = io.BytesIO()
    bytes_file_obj.write(response.content)
    bytes_file_obj.seek(0)
    wb = load_workbook(filename=bytes_file_obj,data_only=True)
    ws = wb[data['sheet_name']]
    ws_range=ws[data['range_from']+str(ws.min_row):data['range_to']+str(ws.max_row)]
    #ws_range=ws[data['range_from']+'1':data['range_to']+'9']
    date_mask_col=[]
    date_mask_format=[]
    for item in data['SourcePossibleDateMask']:
        if "/-/" in item:
            sub_items = item.split(" /-/ ")
            date_mask_col.append(sub_items[1])
            date_mask_format.append(map_csharp_to_python_format(sub_items[0]))
            
    cols=[]   
    for cell in ws_range[0]:
        cols.append(cell.value)
    
    data_rows = []
    for row in ws_range[1:]:
        data_cols = []
        for cell in range(len(row)):
            if isinstance(row[cell].value, datetime):
                data_cols.append(row[cell].value.strftime('%Y-%m-%d %H:%M:%S'))
            elif cols[cell] in  date_mask_col:
                for i in range(len(date_mask_col)):
                    print(date_mask_format[i])
                    if cols[cell] == date_mask_col[i]:
                        datetime_object = datetime.strptime(row[cell].value, date_mask_format[i])
                       
                        data_cols.append(datetime_object.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                data_cols.append(row[cell].value)
        data_rows.append(data_cols)
    keys = cols

    data_dict = [dict(zip(keys, row)) for row in data_rows[0:]]

    json_data = json.dumps(data_dict,indent=4,default=str, ensure_ascii=False)

    return(json.loads(json_data))


@router.post('/remove_object_and_convert_to_lowkey')
async def remove_object_and_convert_to_lowkey(data: Union[List[dict],dict]=Body(...)):
    if isinstance(data, (list)):
        for item in data:
            for key in list(item.keys()):
                if isinstance(item[key], (dict, list)):
                    item.pop(key)
                else:
                    item[unidecode(key.lower().replace(' ', '_'))] = item.pop(key)
    else:
        for key in list(data.keys()):
            if isinstance(data[key], (dict, list)):
                data.pop(key)
            else:
                data[unidecode(key.lower().replace(' ', '_'))] = data.pop(key)
        data=[data]
    
        
    json_str = json.dumps(data)
    return(json.loads(json_str))

@router.post('/convert_id_to_lowkey')
async def convert_id_to_lowkey(data: List[dict]=Body(...),index:str= Header(default=None)):
    index=index.split(", ")
    for item in data:
        for indexname in index:
            if indexname in item:
                item[indexname.lower()] = item.pop(indexname)
        # item = dict(map(lambda item: (unidecode(item[0]), item[1]), item.items()))

    json_str = json.dumps(data)
    return(json.loads(json_str))

@router.post('/elastic_transit_flatten_array')
async def elastic_transit_flatten_array(data: Union[dict,list]=Body(...),
                                        index:str= Header(default=None)):
    
    if isinstance(data, (list)):
       
        df = pl.from_records(data)

    # Select "_source" column
        df = df.select(["_source"])
        df=pl.DataFrame(df["_source"]).unnest('_source')
       
    # Flatten "_source" colum
        flatten_array=df.to_dicts()
        
    else:
        flatten_array.extend([data["_source"]])
    json_str = json.dumps(flatten_array)
    return(json.loads(json_str))

@router.post("/run-script")
async def run_script(data: dict=Body(...), index:str= Header(default=None)):
    script = data['python_script']
    variable_json=None
    variable_string=""
    if 'variable_json' in data:
        variable_json = json.loads(json.dumps(data['variable_json']))
    if 'variable_string' in data:
        variable_string=data['variable_string'  ]
    command = ['python', '-c', script,variable_string]
    result = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    output, error = result.communicate(input=json.dumps(variable_json))
    return {"output": output, "error": error}
@router.post("/run_python_script")
async def run_python_script(data: dict=Body(...), index:str= Header(default=None)):
    script = base64.b64decode(data['python_script'])
    variable_json=None
    variable_string=""
    if 'variable_json' in data:
        variable_json = json.loads(json.dumps(data['variable_json']))
    if 'variable_string' in data:
        variable_string=data['variable_string'  ]
    command = ['python', '-c', script,variable_string]
    result = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    output, error = result.communicate(input=json.dumps(variable_json))
    return {"output": output, "error": error}

@router.post("/upload_avro")
async def upload_avro_file(avro_file: bytes=Body(...),
                           fileName:str= Header(default=None),
                           TableName:str= Header(default=None),
                           serviceAccountjson:str= Header(default=None)):
    try:
        with open(fileName, "wb") as file:
            file.write(avro_file)
        file.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving Avro data: {str(e)}")  
    try:
        service_account_json = json.loads(serviceAccountjson)
        google_credentials = service_account.Credentials.from_service_account_info(
            service_account_json)
        client = bigquery.Client(credentials=google_credentials)
        job_config = bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.AVRO
            ,create_disposition="CREATE_IF_NEEDED"
            ,write_disposition="WRITE_TRUNCATE"
        )
        with open(fileName, "rb") as source_file:
            job = client.load_table_from_file(source_file, TableName, job_config=job_config)
        job.result()
        return {"detail": "Success"} 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Put data Error: {str(e)}")  
    finally:
        try:
            os.remove(fileName)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error deleting file: {str(e)}")  




    
